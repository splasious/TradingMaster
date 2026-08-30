"""Excel export (PRD section 7): single symbol, a full watchlist (one sheet
per symbol), or everything tracked. Every workbook gets a summary sheet
first with real metadata -- source, symbol, timeframe, date range, and
export timestamp -- not just raw rows.
"""

import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.backfill_platform import BfOhlcvBar, BfSymbol, BfWatchlistItem

_HEADER = ["Timestamp (UTC)", "Open", "High", "Low", "Close", "Volume"]


def _sheet_name(symbol: BfSymbol) -> str:
    # Excel sheet names: max 31 chars, no []:*?/\\
    raw = f"{symbol.source}_{symbol.symbol}"
    cleaned = "".join(c for c in raw if c not in '[]:*?/\\')
    return cleaned[:31] or "sheet"


def _write_bars_sheet(wb: Workbook, symbol: BfSymbol, bars: list[BfOhlcvBar]) -> None:
    ws = wb.create_sheet(title=_sheet_name(symbol))
    ws.append(_HEADER)
    for bar in sorted(bars, key=lambda b: b.ts):
        ws.append([bar.ts.strftime("%Y-%m-%d %H:%M:%S"), bar.open, bar.high, bar.low, bar.close, bar.volume])
    for i, _ in enumerate(_HEADER, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 16


def _write_summary_sheet(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Source", "Symbol", "Name", "Timeframe", "Bars", "Earliest", "Latest", "Exported (UTC)"])
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    for row in rows:
        ws.append([
            row["source"], row["symbol"], row["display_name"], row["timeframe"], row["bar_count"],
            row["earliest"].strftime("%Y-%m-%d %H:%M:%S") if row["earliest"] else "--",
            row["latest"].strftime("%Y-%m-%d %H:%M:%S") if row["latest"] else "--",
            now,
        ])
    for i in range(1, 9):
        ws.column_dimensions[get_column_letter(i)].width = 18


async def _bars_for_symbol(db: AsyncSession, symbol_id, timeframe: str | None = None) -> list[BfOhlcvBar]:
    stmt = select(BfOhlcvBar).where(BfOhlcvBar.symbol_id == symbol_id)
    if timeframe:
        stmt = stmt.where(BfOhlcvBar.timeframe == timeframe)
    return list((await db.execute(stmt.order_by(BfOhlcvBar.ts))).scalars().all())


async def export_symbol(db: AsyncSession, symbol: BfSymbol, timeframe: str) -> bytes:
    bars = await _bars_for_symbol(db, symbol.id, timeframe)
    wb = Workbook()
    _write_summary_sheet(wb, [{
        "source": symbol.source, "symbol": symbol.symbol, "display_name": symbol.display_name, "timeframe": timeframe,
        "bar_count": len(bars), "earliest": bars[0].ts if bars else None, "latest": bars[-1].ts if bars else None,
    }])
    _write_bars_sheet(wb, symbol, bars)
    return _to_bytes(wb)


async def export_watchlist(db: AsyncSession, watchlist_id, timeframe: str | None) -> bytes:
    items = (
        await db.execute(select(BfWatchlistItem).where(BfWatchlistItem.watchlist_id == watchlist_id))
    ).scalars().all()
    wb = Workbook()
    summary_rows = []
    symbol_bars: list[tuple[BfSymbol, list[BfOhlcvBar]]] = []
    for item in items:
        symbol = await db.get(BfSymbol, item.symbol_id)
        if symbol is None:
            continue
        bars = await _bars_for_symbol(db, symbol.id, timeframe)
        symbol_bars.append((symbol, bars))
        summary_rows.append({
            "source": symbol.source, "symbol": symbol.symbol, "display_name": symbol.display_name,
            "timeframe": timeframe or "all", "bar_count": len(bars),
            "earliest": bars[0].ts if bars else None, "latest": bars[-1].ts if bars else None,
        })
    _write_summary_sheet(wb, summary_rows)
    for symbol, bars in symbol_bars:
        _write_bars_sheet(wb, symbol, bars)
    return _to_bytes(wb)


async def export_all(db: AsyncSession) -> bytes:
    symbols = list((await db.execute(select(BfSymbol))).scalars().all())
    wb = Workbook()
    summary_rows = []
    symbol_bars: list[tuple[BfSymbol, list[BfOhlcvBar]]] = []
    for symbol in symbols:
        bars = await _bars_for_symbol(db, symbol.id)
        symbol_bars.append((symbol, bars))
        summary_rows.append({
            "source": symbol.source, "symbol": symbol.symbol, "display_name": symbol.display_name, "timeframe": "all",
            "bar_count": len(bars), "earliest": bars[0].ts if bars else None, "latest": bars[-1].ts if bars else None,
        })
    _write_summary_sheet(wb, summary_rows)
    for symbol, bars in symbol_bars:
        _write_bars_sheet(wb, symbol, bars)
    return _to_bytes(wb)


def _to_bytes(wb: Workbook) -> bytes:
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
